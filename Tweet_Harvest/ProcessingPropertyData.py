__author__ = 'tuhung-te'
import DataLoader
import openpyxl
import DBManager
import Settings as config
import time
class ProcessingProertyData(object):

    def __init__(self):
        self.a = "a"


    def auctionHouse(self):
        data = DataLoader.DataLoader().loadCSV("/Users/tuhung-te/desktop/Final/Property_Level_Data_of_Auctioned_Properties_Houses_2011_for_VIC/data6755821851114529791.csv")
        for row in data:
            print(row)

    def HousingIndex(self):
        worksheet = openpyxl.load_workbook("/Users/tuhung-te/desktop/Final/Data/House_Pricing_Index.xlsx")
        sheet = worksheet.get_sheet_by_name("Data1")
        db = DBManager.DBManager()
        db.DBalive()
        for row in range(11,sheet.get_highest_row()+1):
            index = sheet["C"+str(row)].value
            date = str(sheet["A"+str(row)].value)
            db.transaction({"index":index,"date":date}
                           ,config.DBMapping.get("houseingindex",lambda: "nothing")()["DB_Name"])

    def houseing_Price_To_Income_Ratio(self):
        worksheet = openpyxl.load_workbook("/Users/tuhung-te/desktop/Final/Data/House_Price_Income_Ratio.xlsx")
        sheet = worksheet.get_sheet_by_name("Sheet1")
        db = DBManager.DBManager()
        db.DBalive()
        for row in range(11,sheet.get_highest_row()+1):
            date = str(sheet["A"+str(row)].value)
            houseprice =sheet["B"+str(row)].value * 1000
            weekly_income = sheet["D"+str(row)].value *42
            ratio = houseprice/ weekly_income
            db.transaction({"ratio":ratio,"date":date, "attribute":{"Annual_income":weekly_income,"houseprice":houseprice}}
                                                ,config.DBMapping.get("housepriceincome",lambda: "nothing")()["DB_Name"])

    def housePriceWithRenterPayment(self):
        worksheet = openpyxl.load_workbook("/Users/tuhung-te/desktop/Final/Data/Gross_Renter_Payment.xlsx")
        sheet = worksheet.get_sheet_by_name("Sheet1")
        db = DBManager.DBManager()
        db.DBalive()
        for row in range(2,sheet.get_highest_row()+1):
            date = str(sheet["A"+str(row)].value)
            renter_payment = sheet["B"+str(row)].value * 42
            house_price =sheet["C"+str(row)].value * 1000
            house_with_renter = house_price/renter_payment
            db.transaction({"ratio":house_with_renter,"date":date, "attribute":{"renter_payment":renter_payment,"houseprice":house_price}}
                                                ,config.DBMapping.get("housepricewithrenterpayment",lambda: "nothing")()["DB_Name"])

    def mortgageWithGDP(self):
        worksheet = openpyxl.load_workbook("/Users/tuhung-te/desktop/Final/Data/MortgageDebt_GDP.xlsx")
        sheet = worksheet.get_sheet_by_name("FRED Graph")
        db = DBManager.DBManager()
        db.DBalive()
        for row in range(12,sheet.get_highest_row()+1):
            date = str(sheet["A"+str(row)].value)
            mortgageWithGDP = sheet["B"+str(row)].value
            db.transaction({"mortgageWithGDP":mortgageWithGDP,"date":date}
                                                ,config.DBMapping.get("mortgagewithgdp",lambda: "nothing")()["DB_Name"])


if __name__ == '__main__':
    ProcessingProertyData().mortgageWithGDP();